"""The pull-sheet workbook: a FILLED COPY of the user's own template.

The user keeps a Google Sheets pull sheet (exported to xlsx and shipped,
scrubbed, as static/templates/pull-sheet-template.xlsx). Its shape is the
contract this module writes to, nothing is invented here:

  Pull Sheet   positions SIDE BY SIDE in 6-column blocks from column A
               (Cable Type | Length | Qty | Label | Notes | spacer): blocks
               at A, G, M, S, Y, AE. Row 5 holds the position name (the
               first two cells merged), row 6 the headers, rows 7-36 the
               data (30 rows). The fourth block (S) is the tan "NEW
               POSITION" copy-me block. B2 show, B3 engineer, E2 date, E3
               rev.
  GEAR LIST    A4:A… cable types, B4:B… lengths ("6'", "25'", "EA") - the
               source of both dropdowns on the Pull Sheet.
  Spares, TOTALS, calc (hidden)
               formula-built: calc scans the six blocks × rows 7-66 through
               INDIRECT() string literals, so SIX positions of SIXTY rows is
               the hard limit of the sheet, and the rows must be written IN
               PLACE - inserting rows would not move those literals.

openpyxl round-trips every formula (array formulas included), both data
validations, the merged cells, the freeze panes and the hidden calc tab;
what it drops is six empty drawing stubs and an empty threaded-comments
list, neither of which carries anything.

Three entry points:
  gear_list()                the GEAR LIST vocabulary (types, lengths) the
                             in-app editor's pickers draw from
  scrub_template(src, dst)   build the shipped template from a copy of the
                             user's sheet (data rows, show, engineer, rev
                             and date cleared; positions renamed)
  build_workbook(pull_list, meta) -> (bytes, warnings)
                             fill a copy of the template with the client's
                             pull list (app-pull-list.js buildPullList)
"""
import copy
import datetime
import io
import os
import sys

BASE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(BASE_DIR, 'static', 'templates', 'pull-sheet-template.xlsx')

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

SHEET = 'Pull Sheet'
GEAR = 'GEAR LIST'
OVERFLOW = 'Overflow'

# The six blocks' first columns (1-based): A, G, M, S, Y, AE.
BLOCK_COLS = [1, 7, 13, 19, 25, 31]
BLOCK_WIDTH = 5            # Cable Type, Length, Qty, Label, Notes
BLOCK_STRIDE = 6           # plus the spacer column
TITLE_ROW = 5
HEADER_ROW = 6
FIRST_DATA_ROW = 7
LAST_DATA_ROW = 36         # the template's 30 rows
MAX_DATA_ROW = 66          # the calc tab scans to here and no further
MAX_POSITIONS = len(BLOCK_COLS)
TEMPLATE_BLOCK = 3         # index of the tan NEW POSITION block (S)
TEMPLATE_TITLE = 'NEW POSITION'
TEMPLATE_NOTE = '< copy to add a position >'
HEADERS = ['Cable Type', 'Length', 'Qty', 'Label', 'Notes']

GEAR_TYPE_COL, GEAR_LENGTH_COL, GEAR_FIRST_ROW = 1, 2, 4
GEAR_TYPE_LAST_ROW, GEAR_LENGTH_LAST_ROW = 200, 60   # the validations' ranges


def _openpyxl():
    import openpyxl  # noqa: F401  (imported lazily so the app boots without it)
    return openpyxl


def _col_letter(col):
    from openpyxl.utils import get_column_letter
    return get_column_letter(col)


def _style_of(cell):
    """A cell's style, detached from the cell (a snapshot must not follow
    later edits to the cell it was taken from)."""
    if not cell.has_style:
        return None
    return {
        'font': copy.copy(cell.font), 'fill': copy.copy(cell.fill),
        'border': copy.copy(cell.border), 'alignment': copy.copy(cell.alignment),
        'number_format': cell.number_format, 'protection': copy.copy(cell.protection),
    }


def _apply_style(style, dst):
    if not style:
        return
    dst.font = copy.copy(style['font'])
    dst.fill = copy.copy(style['fill'])
    dst.border = copy.copy(style['border'])
    dst.alignment = copy.copy(style['alignment'])
    dst.number_format = style['number_format']
    dst.protection = copy.copy(style['protection'])


def _copy_style(src, dst):
    _apply_style(_style_of(src), dst)


def _block_cells(ws, block, first_row, last_row):
    col0 = BLOCK_COLS[block]
    for r in range(first_row, last_row + 1):
        for c in range(col0, col0 + BLOCK_WIDTH):
            yield ws.cell(r, c)


# ── the shipped template ─────────────────────────────────────────────────

def scrub_template(src_path, dst_path):
    """Write the shipped template from a copy of the user's sheet: the three
    named positions' data rows cleared (styles kept), the positions renamed
    POSITION 1-3, show / engineer / rev / date blanked. Everything else -
    every tab, formula, validation, merge, freeze pane, the hidden calc
    state and the GEAR LIST vocabulary - is left exactly as it was. Running
    it over the shipped template again changes nothing."""
    wb = _openpyxl().load_workbook(src_path)
    ws = wb[SHEET]
    for block in range(TEMPLATE_BLOCK):
        for cell in _block_cells(ws, block, FIRST_DATA_ROW, LAST_DATA_ROW):
            cell.value = None
        ws.cell(TITLE_ROW, BLOCK_COLS[block]).value = f'POSITION {block + 1}'
    for ref in ('B2', 'B3', 'E2', 'E3'):
        ws[ref].value = None
    wb.save(dst_path)


# ── the filled copy ──────────────────────────────────────────────────────

def _positions(pull_list):
    out = []
    for p in (pull_list or {}).get('positions') or []:
        if not isinstance(p, dict):
            continue
        rows = []
        for r in p.get('rows') or []:
            if not isinstance(r, dict) or not str(r.get('type') or '').strip():
                continue
            try:
                qty = int(round(float(r.get('qty') or 0)))
            except (TypeError, ValueError):
                qty = 0
            if qty <= 0:
                continue
            rows.append({
                'type': str(r.get('type')).strip(),
                'length': str(r.get('length') or '').strip(),
                'qty': qty,
                'label': str(r.get('label') or ''),
                'notes': str(r.get('notes') or ''),
            })
        out.append({'name': str(p.get('name') or '').strip() or 'Position', 'rows': rows})
    return out


def _gear_column(ws, col, last_row):
    """(values, next free row) for one GEAR LIST column."""
    values = []
    free = None
    for r in range(GEAR_FIRST_ROW, last_row + 1):
        v = ws.cell(r, col).value
        if v is None or str(v).strip() == '':
            if free is None:
                free = r
            continue
        values.append(str(v).strip())
    return values, free


def _add_gear(ws, col, last_row, wanted, warnings, what):
    """Append every `wanted` entry the column lacks, styled like the row
    above it, so the dropdowns accept what the sheet lists."""
    have, free = _gear_column(ws, col, last_row)
    have_set = set(have)
    for v in wanted:
        if not v or v in have_set:
            continue
        if free is None or free > last_row:
            warnings.append(f'GEAR LIST is full: "{v}" was not added to the {what} dropdown.')
            continue
        cell = ws.cell(free, col)
        cell.value = v
        _copy_style(ws.cell(free - 1, col), cell)
        have_set.add(v)
        free += 1


def _extend_validations(ws, last_row, used_blocks):
    """Re-aim both dropdown validations at rows 7..last_row of every block
    in use (the template names A/G/M/S only, and rows to 36)."""
    from openpyxl.worksheet.cell_range import MultiCellRange
    for dv in ws.data_validations.dataValidation:
        cols = sorted({rng.min_col for rng in dv.sqref.ranges})
        if not cols:
            continue
        offset = cols[0] - BLOCK_COLS[0]      # 0 for the type column, 1 for length
        if offset not in (0, 1):
            continue
        refs = []
        for block in range(MAX_POSITIONS):
            if block not in used_blocks and block > TEMPLATE_BLOCK:
                continue
            col = _col_letter(BLOCK_COLS[block] + offset)
            refs.append(f'{col}{FIRST_DATA_ROW}:{col}{last_row}')
        dv.sqref = MultiCellRange(' '.join(refs))


def _extend_rows(ws, last_row, blocks):
    """Carry the data-row styling down to `last_row` for the given blocks -
    the in-place equivalent of the READ ME's "insert rows to all positions"
    (the calc tab already scans to row 66)."""
    if last_row <= LAST_DATA_ROW:
        return
    for block in blocks:
        col0 = BLOCK_COLS[block]
        for r in range(LAST_DATA_ROW + 1, last_row + 1):
            for c in range(col0, col0 + BLOCK_WIDTH):
                _copy_style(ws.cell(LAST_DATA_ROW, c), ws.cell(r, c))
            if ws.row_dimensions[LAST_DATA_ROW].height:
                ws.row_dimensions[r].height = ws.row_dimensions[LAST_DATA_ROW].height


def _snapshot_block(ws, block, last_row):
    """Styles, values, merges and widths of one block, for cloning."""
    col0 = BLOCK_COLS[block]
    cells = {}
    for r in range(TITLE_ROW, last_row + 1):
        for c in range(col0, col0 + BLOCK_STRIDE):
            src = ws.cell(r, c)
            cells[(r, c - col0)] = (src.value, _style_of(src))
    merges = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == TITLE_ROW and col0 <= rng.min_col < col0 + BLOCK_STRIDE:
            merges.append((rng.min_col - col0, rng.max_col - col0))
    widths = {}
    for c in range(col0, col0 + BLOCK_STRIDE):
        dim = ws.column_dimensions.get(_col_letter(c))
        widths[c - col0] = dim.width if dim is not None else None
    return {'cells': cells, 'merges': merges, 'widths': widths}


def _clone_block(ws, snapshot, block, last_row, with_values):
    col0 = BLOCK_COLS[block]
    for (r, dc), (value, style) in snapshot['cells'].items():
        if r > last_row:
            continue
        dst = ws.cell(r, col0 + dc)
        _apply_style(style, dst)
        if with_values:
            dst.value = value
    for (dc0, dc1) in snapshot['merges']:
        ws.merge_cells(start_row=TITLE_ROW, start_column=col0 + dc0,
                       end_row=TITLE_ROW, end_column=col0 + dc1)
    for dc, width in snapshot['widths'].items():
        if width:
            ws.column_dimensions[_col_letter(col0 + dc)].width = width
    for c in range(col0, col0 + BLOCK_WIDTH):
        ws.cell(HEADER_ROW, c).value = HEADERS[c - col0]


def _style_title(ws, block, used_snapshot):
    """A used block's title row reads like the first position's (blue,
    white), never the tan copy-me look."""
    col0 = BLOCK_COLS[block]
    for dc in range(BLOCK_WIDTH):
        _, style = used_snapshot['cells'][(TITLE_ROW, dc)]
        _apply_style(style, ws.cell(TITLE_ROW, col0 + dc))
    ws.cell(TITLE_ROW, col0 + 2).value = None      # the copy-me note's cell


def build_workbook(pull_list, meta=None, template_path=TEMPLATE_PATH):
    """Fill a copy of the template. Returns (xlsx bytes, warnings)."""
    openpyxl = _openpyxl()
    meta = meta or {}
    warnings = []
    wb = openpyxl.load_workbook(template_path)
    ws = wb[SHEET]
    gear = wb[GEAR]

    # Header cells: show, engineer, rev, date (a real date so the sheet's
    # own formatting applies; the display text is what the client shows).
    ws['B2'].value = str(meta.get('project_name') or '').strip() or None
    ws['B3'].value = str(meta.get('engineer') or '').strip() or None
    rev = meta.get('rev')
    ws['E3'].value = (str(rev).strip() or None) if rev is not None else None
    date_iso = str(meta.get('date_iso') or '').strip()
    try:
        ws['E2'].value = datetime.date.fromisoformat(date_iso)
        ws['E2'].number_format = 'mmm d, yyyy'
    except ValueError:
        ws['E2'].value = str(meta.get('date') or '').strip() or None

    positions = _positions(pull_list)
    placed = positions[:MAX_POSITIONS]
    overflow_positions = positions[MAX_POSITIONS:]
    used_blocks = set(range(len(placed)))

    # Rows per block: the template's 30, grown in place to the calc tab's
    # 60 when a position needs more; anything past 60 overflows.
    max_rows = max([len(p['rows']) for p in placed] or [0])
    room = MAX_DATA_ROW - FIRST_DATA_ROW + 1
    last_row = min(MAX_DATA_ROW, max(LAST_DATA_ROW, FIRST_DATA_ROW + max_rows - 1))

    first_snapshot = _snapshot_block(ws, 0, LAST_DATA_ROW)
    template_snapshot = _snapshot_block(ws, TEMPLATE_BLOCK, LAST_DATA_ROW)

    # Blocks Y and AE carry no styling in the template: clone the S block's.
    for block in used_blocks:
        if block > TEMPLATE_BLOCK:
            _clone_block(ws, template_snapshot, block, LAST_DATA_ROW, with_values=False)
    # A used S block (or a used Y/AE) is a position, not the copy-me block.
    for block in used_blocks:
        if block >= TEMPLATE_BLOCK:
            _style_title(ws, block, first_snapshot)
    # The copy-me block moves to the next free block, if there is one.
    next_free = len(placed)
    if TEMPLATE_BLOCK in used_blocks and next_free < MAX_POSITIONS:
        _clone_block(ws, template_snapshot, next_free, LAST_DATA_ROW, with_values=True)
        col0 = BLOCK_COLS[next_free]
        ws.cell(TITLE_ROW, col0).value = TEMPLATE_TITLE
        ws.cell(TITLE_ROW, col0 + 2).value = TEMPLATE_NOTE
        c0, c1 = _col_letter(col0), _col_letter(col0 + BLOCK_WIDTH - 1)
        ref = f"'{SHEET}'!${c0}${TITLE_ROW}:${c1}${LAST_DATA_ROW}"
        from openpyxl.workbook.defined_name import DefinedName
        try:
            del wb.defined_names['NEW_BLOCK_TEMPLATE']
        except KeyError:
            pass
        wb.defined_names['NEW_BLOCK_TEMPLATE'] = DefinedName('NEW_BLOCK_TEMPLATE', attr_text=ref)
        used_blocks_for_rows = used_blocks | {next_free}
    else:
        used_blocks_for_rows = used_blocks | ({TEMPLATE_BLOCK} if TEMPLATE_BLOCK not in used_blocks else set())

    _extend_rows(ws, last_row, sorted(used_blocks_for_rows))
    _extend_validations(ws, last_row, used_blocks_for_rows)

    # The rows themselves.
    overflow_rows = []       # (position name, row)
    types, lengths = [], []
    for block, pos in enumerate(placed):
        col0 = BLOCK_COLS[block]
        ws.cell(TITLE_ROW, col0).value = pos['name']
        for i, r in enumerate(pos['rows']):
            if i >= room:
                overflow_rows.append((pos['name'], r))
                continue
            row = FIRST_DATA_ROW + i
            ws.cell(row, col0).value = r['type']
            ws.cell(row, col0 + 1).value = r['length'] or None
            ws.cell(row, col0 + 2).value = r['qty']
            ws.cell(row, col0 + 3).value = r['label'] or None
            ws.cell(row, col0 + 4).value = r['notes'] or None
            types.append(r['type'])
            if r['length']:
                lengths.append(r['length'])
    for pos in overflow_positions:
        for r in pos['rows']:
            overflow_rows.append((pos['name'], r))

    # Anything the GEAR LIST lacks, so the dropdowns accept what was written.
    _add_gear(gear, GEAR_TYPE_COL, GEAR_TYPE_LAST_ROW, list(dict.fromkeys(types)), warnings, 'Cable Type')
    _add_gear(gear, GEAR_LENGTH_COL, GEAR_LENGTH_LAST_ROW, list(dict.fromkeys(lengths)), warnings, 'Length')

    if overflow_positions:
        names = ', '.join(p['name'] for p in overflow_positions)
        warnings.append(
            f'The sheet holds {MAX_POSITIONS} positions; {names} went to the Overflow tab '
            f'and are not in TOTALS or Spares.')
    if any(i >= room for p in placed for i in range(len(p['rows']))):
        warnings.append(
            f'A position has more than {room} rows; the rest went to the Overflow tab '
            f'and are not in TOTALS or Spares.')
    if overflow_rows:
        ov = wb.create_sheet(OVERFLOW)
        ov.append(['Position'] + HEADERS)
        for name, r in overflow_rows:
            ov.append([name, r['type'], r['length'] or None, r['qty'], r['label'] or None,
                       r['notes'] or None])
        ov.freeze_panes = 'A2'
        for col, width in zip('ABCDEF', (22, 26, 8, 6, 20, 20)):
            ov.column_dimensions[col].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), warnings


def gear_list(template_path=TEMPLATE_PATH):
    """The GEAR LIST vocabulary as the template carries it: the cable types
    (column A) and the lengths (column B) both Pull Sheet dropdowns draw
    from. The in-app pull-sheet editor feeds its own pickers from this, so
    the two never drift - the template is the one source."""
    wb = _openpyxl().load_workbook(template_path, read_only=True)
    ws = wb[GEAR]
    types, lengths = [], []
    for r in range(GEAR_FIRST_ROW, GEAR_TYPE_LAST_ROW + 1):
        v = ws.cell(r, GEAR_TYPE_COL).value
        if v is not None and str(v).strip():
            types.append(str(v).strip())
        if r <= GEAR_LENGTH_LAST_ROW:
            v = ws.cell(r, GEAR_LENGTH_COL).value
            if v is not None and str(v).strip():
                lengths.append(str(v).strip())
    wb.close()
    return {'types': list(dict.fromkeys(types)), 'lengths': list(dict.fromkeys(lengths))}


def safe_filename(name):
    cleaned = ''.join('_' if ch in '\\/:*?"<>|' or ord(ch) < 32 else ch for ch in str(name or ''))
    cleaned = cleaned.strip(' .')
    return cleaned or 'Project'


if __name__ == '__main__':
    # python pull_sheet.py scrub <source.xlsx> <destination.xlsx>
    if len(sys.argv) == 4 and sys.argv[1] == 'scrub':
        scrub_template(sys.argv[2], sys.argv[3])
        print(f'wrote {sys.argv[3]}')
    else:
        print(__doc__)
        sys.exit(2)
